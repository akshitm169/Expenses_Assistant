from tkinter import *
from tkinter import ttk
from tkinter import  messagebox as mb
from tkinter.ttk import Notebook
import matplotlib.pyplot as plt
import tkinter.filedialog as f
import openpyxl
from datetime import *
from calendar import monthrange
#from mpl_toolkits.mplot3d import Axes3D
#import numpy as np
MONTHS_LIST=['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun', 'Jul','Aug', 'Sep', 'Oct', 'Nov', 'Dec']
Cat_List=['Entertainment','Education','Shopping','Personal Care','Healthcare','Kids','Food & Dining','Investments','Bills','Transport','Others']
IDict={'Jan':0, 'Feb':0,'Mar':0,'Apr':0,'May':0,'Jun':0,'Jul':0,'Aug':0,'Sep':0,'Oct':0,'Nov':0,'Dec':0}
EDict={'Jan':0, 'Feb':0,'Mar':0,'Apr':0,'May':0,'Jun':0,'Jul':0,'Aug':0,'Sep':0,'Oct':0,'Nov':0,'Dec':0}
CatDict={'Entertainment':0,'Education':0,'Shopping':0,'Personal Care':0,'Healthcare':0,'Kids':0,'Food & Dining':0,'Investments':0,'Bills':0,'Transport':0,'Others':0}
rows, cols = (12,11)
matrix = [[0]*cols]*rows

left = [1,3,5,7,9,11,13,15,17,19,21,23]
temp=0
path=None
now=  datetime.now()
day=now.day
month=now.month
year=now.year




def Addincome():
    a=SMONTH.get()
    b=Source_of_Income.get()
    c=Income.get()
    income_data = [a, b,c]
    TV1.insert('', 'end', values=income_data)
    IDict[a]=IDict[a]+c
    if plt.fignum_exists(1):
        f = plt.figure(1)
        plt.bar(left, list(IDict.values()), tick_label=list(IDict.keys()), width=0.8, color=['red', 'green'])
        plt.xlabel('Months')
        plt.ylabel('Income')
        plt.title('Month vs Income')
        f.show()





def Deleteincome():
    a = TV1.item(TV1.selection())['values'][0]
    c = TV1.item(TV1.selection())['values'][2]
    IDict[a] = IDict[a] - c
    if plt.fignum_exists(1):
        f = plt.figure(1)
        plt.bar(left, list(IDict.values()), tick_label=list(IDict.keys()), width=0.8, color=['red', 'green'])
        plt.xlabel('Months')
        plt.ylabel('Income')
        plt.title('Month vs Income')
        plt.clf()
        f = plt.figure(1)
        plt.bar(left, list(IDict.values()), tick_label=list(IDict.keys()), width=0.8, color=['red', 'green'])
        plt.xlabel('Months')
        plt.ylabel('Income')
        plt.title('Month vs Income')
        f.show()

    TV1.delete(TV1.selection())


def MI():
    f = plt.figure(1)
    plt.bar(left, list(IDict.values()), tick_label=list(IDict.keys()), width=0.8, color=['red', 'green'])
    plt.xlabel('Months')
    plt.ylabel('Income')
    plt.title('Month vs Income')
    f.show()



def clear_income_data():
    TV1.delete(*TV1.get_children())
    global IDict
    IDict = {'Jan': 0, 'Feb': 0, 'Mar': 0, 'Apr': 0, 'May': 0, 'Jun': 0, 'Jul': 0, 'Aug': 0, 'Sep': 0, 'Oct': 0,
             'Nov': 0, 'Dec': 0}
    if plt.fignum_exists(1):
        f = plt.figure(1)
        plt.bar(left, list(IDict.values()), tick_label=list(IDict.keys()), width=0.8, color=['red', 'green'])
        plt.xlabel('Months')
        plt.ylabel('Income')
        plt.title('Month vs Income')
        plt.clf()
        f = plt.figure(1)
        plt.bar(left, list(IDict.values()), tick_label=list(IDict.keys()), width=0.8, color=['red', 'green'])
        plt.xlabel('Months')
        plt.ylabel('Income')
        plt.title('Month vs Income')
        f.show()



###############################################################



def Addexpense():
    a = SMONTH2.get()
    b = ECategory.get()
    c = Expense.get()
    expense_data = [a, b, c]
    TV2.insert('', 'end', values=expense_data)
    EDict[a] = EDict[a] + c
    CatDict[b]=CatDict[b] + c
    if temp==1:
        wb_obj = openpyxl.load_workbook(path)
        sheet_obj = wb_obj.active
        max_row = sheet_obj.max_row
        if (sheet_obj.cell(row=1, column=1).value == None):
            max_row = 0
        sheet_obj.cell(row=max_row+1,column=1).value=a
        sheet_obj.cell(row=max_row+1, column=2).value =b
        sheet_obj.cell(row=max_row+1, column=3).value =c
        wb_obj.save(path)



    if plt.fignum_exists(2):
        g = plt.figure(2)
        plt.bar(left, list(EDict.values()), tick_label=list(EDict.keys()), width=0.8, color=['red', 'green'])
        plt.xlabel('Months')
        plt.ylabel('Expense')
        plt.title('Month vs Expense')
        g.show()

    if plt.fignum_exists(3):
        h = plt.figure(3)
        patches, texts = plt.pie(list(CatDict.values()), shadow=True, startangle=90)
        plt.legend(patches, list(CatDict.keys()), loc="best")
        plt.axis('equal')
        plt.tight_layout()
        h.show()

    matrix[MONTHS_LIST.index(a)][Cat_List.index(b)]=matrix[MONTHS_LIST.index(a)][Cat_List.index(b)]+c




def Deleteexpense():
    a = TV2.item(TV2.selection())['values'][0]
    b= TV2.item(TV2.selection())['values'][1]
    c = TV2.item(TV2.selection())['values'][2]
    EDict[a] = EDict[a] - c
    CatDict[b] = CatDict[b] - c
    #print(TV2.index(TV2.selection()))

    if temp==1:
        rowoftable=TV2.index(TV2.selection())
        wb_obj = openpyxl.load_workbook(path)
        sheet_obj = wb_obj.active
        sheet_obj.delete_rows(rowoftable+1,amount=1)
        wb_obj.save(path)


    if plt.fignum_exists(2):
        g = plt.figure(2)
        plt.bar(left, list(EDict.values()), tick_label=list(EDict.keys()), width=0.8, color=['red', 'green'])
        plt.xlabel('Months')
        plt.ylabel('Expense')
        plt.title('Month vs Expense')
        plt.clf()
        g = plt.figure(2)
        plt.bar(left, list(EDict.values()), tick_label=list(EDict.keys()), width=0.8, color=['red', 'green'])
        plt.xlabel('Months')
        plt.ylabel('Expense')
        plt.title('Month vs Expense')
        g.show()


    TV2.delete(TV2.selection())

    if plt.fignum_exists(3):
        h = plt.figure(3)
        patches, texts = plt.pie(list(CatDict.values()), shadow=True, startangle=90)
        plt.legend(patches, list(CatDict.keys()), loc="best")
        plt.axis('equal')
        plt.tight_layout()
        plt.clf()
        h = plt.figure(3)
        patches, texts = plt.pie(list(CatDict.values()), shadow=True, startangle=90)
        plt.legend(patches, list(CatDict.keys()), loc="best")
        plt.axis('equal')
        plt.tight_layout()
        h.show()

    matrix[MONTHS_LIST.index(a)][Cat_List.index(b)]=matrix[MONTHS_LIST.index(a)][Cat_List.index(b)]- c



def Show_Cat():
    h=plt.figure(3)
    patches, texts = plt.pie(list(CatDict.values()), shadow=True, startangle=90)
    plt.legend(patches,list(CatDict.keys()) , loc="best")
    plt.axis('equal')
    plt.tight_layout()
    h.show()



def ME():
    g = plt.figure(2)
    plt.bar(left, list(EDict.values()), tick_label=list(EDict.keys()), width=0.8, color=['red', 'green'])
    plt.xlabel('Months')
    plt.ylabel('Expense')
    plt.title('Month vs Expense')
    g.show()




def Show_Over():
    pass
        #TV2.detach(each)
#    for child in TV2.get_children():
#        print(TV2.item(child)["values"])
#    print("fgb")



def Open_Existing():
    global temp
    global path
    temp=1
    path = f.askopenfilename()
    wb_obj = openpyxl.load_workbook(path)
    sheet_obj = wb_obj.active
    max_row = sheet_obj.max_row
    if(sheet_obj.cell(row = 1, column = 1).value==None):
        max_row=0

    # Will give a particular row value
    for i in range(1, max_row + 1):
        a = (sheet_obj.cell(row=i, column=1)).value
        b = (sheet_obj.cell(row=i, column=2)).value
        c = (sheet_obj.cell(row=i, column=3)).value
        expense_data = [a, b, c]
        TV2.insert('', 'end', values=expense_data)
        EDict[a] = EDict[a] + c
        CatDict[b] = CatDict[b] + c
    if plt.fignum_exists(2):
        g = plt.figure(2)
        plt.bar(left, list(EDict.values()), tick_label=list(EDict.keys()), width=0.8, color=['red', 'green'])
        plt.xlabel('Months')
        plt.ylabel('Expense')
        plt.title('Month vs Expense')
        g.show()
    if plt.fignum_exists(3):
        h = plt.figure(3)
        patches, texts = plt.pie(list(CatDict.values()), shadow=True, startangle=90)
        plt.legend(patches, list(CatDict.keys()), loc="best")
        plt.axis('equal')
        plt.tight_layout()
        h.show()



def Open_New():
    global  temp
    global path
    temp =1
    path=f.asksaveasfilename(filetypes=(("Excel files", "*.xlsx"),("All files", "*.*") ))
    wb = openpyxl.Workbook()  # open a workbook
    wb.save(path)



def clear_expense_data():
    TV2.delete(*TV2.get_children())
    global EDict
    global CatDict
    EDict = {'Jan': 0, 'Feb': 0, 'Mar': 0, 'Apr': 0, 'May': 0, 'Jun': 0, 'Jul': 0, 'Aug': 0, 'Sep': 0, 'Oct': 0,
             'Nov': 0, 'Dec': 0}
    CatDict = {'Entertainment': 0, 'Education': 0, 'Shopping': 0, 'Personal Care': 0, 'Healthcare': 0, 'Kids': 0,
               'Food & Dining': 0, 'Investments': 0, 'Bills': 0, 'Transport': 0, 'Others': 0}

    if plt.fignum_exists(2):
        g = plt.figure(2)
        plt.bar(left, list(EDict.values()), tick_label=list(EDict.keys()), width=0.8, color=['red', 'green'])
        plt.xlabel('Months')
        plt.ylabel('Expense')
        plt.title('Month vs Expense')
        plt.clf()
        g = plt.figure(2)
        plt.bar(left, list(EDict.values()), tick_label=list(EDict.keys()), width=0.8, color=['red', 'green'])
        plt.xlabel('Months')
        plt.ylabel('Expense')
        plt.title('Month vs Expense')
        g.show()


    if plt.fignum_exists(3):
        h = plt.figure(3)
        patches, texts = plt.pie(list(CatDict.values()), shadow=True, startangle=90)
        plt.legend(patches, list(CatDict.keys()), loc="best")
        plt.axis('equal')
        plt.tight_layout()
        plt.clf()
        h = plt.figure(3)
        patches, texts = plt.pie(list(CatDict.values()), shadow=True, startangle=90)
        plt.legend(patches, list(CatDict.keys()), loc="best")
        plt.axis('equal')
        plt.tight_layout()
        h.show()


    if temp==1:
        wb_obj = openpyxl.load_workbook(path)
        sheet_obj = wb_obj.active
        max_row = sheet_obj.max_row
        for i in range(0,max_row):
            sheet_obj.delete_rows(1, amount=1)
        wb_obj.save(path)











def Show_Status():
    limitvar=0
    limitvar=MAXExpense.get()
    virtual_exp = (EDict[MONTHS_LIST[month - 1]] * (monthrange(year, month))[1]) / day
    if(EDict[MONTHS_LIST[month-1]]>limitvar):
        mb.showwarning("STATUS","Oops!! You have already crossed the limit..Try to save next month!!")
    elif(EDict[MONTHS_LIST[month-1]]==limitvar):
        mb.showwarning("STATUS", "You have achieved the limit..Try not to spend more!!")

    elif(virtual_exp<=limitvar):
        mb.showinfo("STATUS","Hurray!! According to your expenditure till now it is expected that you will achieve your target!!")
    elif(virtual_exp>limitvar):
        mb.showwarning("STATUS","Oops!! According to your expenditure till now it is expected that you may cross the limit..try to spend less!!")




#CREATING WINDOW
GUI=Tk()
GUI.title('Expense and Income Recorder')
GUI.geometry('820x600')

#CREAING TWO FRAMES
Tab=Notebook(GUI)
F1=Frame(Tab,width=500,height=500)
F2=Frame(Tab,width=500,height=500)
Tab.add(F1, text='Income')
Tab.add(F2, text='Expense')
Tab.pack(fill=BOTH,expand=1)


#INCOME TAB


LMONTH= ttk.Label(F1, text='Month',font=(None,18))
LMONTH.grid(row=0, column=0,padx=5,pady=5,sticky='w')

SMONTH=ttk.Combobox(F1,values=MONTHS_LIST,font=(None,18),width=20,state="readonly")
SMONTH.grid(row=0, column=1,padx=5,pady=5,sticky='w')
SMONTH.current(0)


LSource= ttk.Label(F1, text='Source of Income',font=(None,18))
LSource.grid(row=1, column=0,padx=5,pady=5,sticky='w')

Source_of_Income=StringVar()

ESource= ttk.Entry(F1, textvariable=Source_of_Income,font=(None,18))
ESource.grid(row=1, column=1,padx=5,pady=5,sticky='w')


LIncome= ttk.Label(F1, text='Income',font=(None,18))
LIncome.grid(row=2, column=0,padx=5,pady=5,sticky='w')

Income=IntVar()

EIncome= ttk.Entry(F1, textvariable=Income,font=(None,18))
EIncome.grid(row=2, column=1,padx=5,pady=5,sticky='w')


BF1Add=ttk.Button(F1, text='Add Income',command=Addincome)
BF1Add.grid(row=3, column=1,padx=5,pady=5,sticky='w',ipadx=10,ipady=10)


BF1Delete=ttk.Button(F1, text='Delete Income',command=Deleteincome)
BF1Delete.grid(row=4, column=1,padx=5,pady=5,sticky='w',ipadx=10,ipady=10)

Show_Bar_Button=ttk.Button(F1, text='Month Vs Income',command=MI)
Show_Bar_Button.grid(row=0, column=2,padx=5,pady=5,sticky='w',ipadx=10,ipady=10)


clear_Button=ttk.Button(F1, text='Clear',command=clear_income_data)
clear_Button.grid(row=5, column=3,padx=5,pady=5,sticky='w',ipadx=10,ipady=10)


#---TREE VIEW---
TVList1 =['Month','Source of Income','Income']
TV1 =ttk.Treeview(F1, column=TVList1, show='headings',height=14)
for i in TVList1:
    TV1.heading(i, text=i.title())
TV1.grid(row=5, column=0,padx=5,pady=5,sticky='w',columnspan=3)


###############################################################################

LMONTH2= ttk.Label(F2, text='Month',font=(None,18))
LMONTH2.grid(row=0, column=0,padx=5,pady=5,sticky='w')

SMONTH2=ttk.Combobox(F2,values=MONTHS_LIST,font=(None,18),width=20,state="readonly")
SMONTH2.grid(row=0, column=1,padx=5,pady=5,sticky='w')
SMONTH2.current(0)




LCategory= ttk.Label(F2, text='Category',font=(None,18))
LCategory.grid(row=1, column=0,padx=5,pady=5,sticky='w')


ECategory= ttk.Combobox(F2,values=Cat_List ,font=(None,18),width=20,state="readonly")
ECategory.grid(row=1, column=1,padx=5,pady=5,sticky='w')
ECategory.current(0)


LExpense= ttk.Label(F2, text='Expense',font=(None,18))
LExpense.grid(row=2, column=0,padx=5,pady=5,sticky='w')

Expense=IntVar()

EExpense= ttk.Entry(F2, textvariable=Expense,font=(None,18))
EExpense.grid(row=2, column=1,padx=5,pady=5,sticky='w')



LMAXExpense= ttk.Label(F2, text='Current Month Limit',font=(None,12))
LMAXExpense.grid(row=3, column=2,padx=5,pady=5,sticky='w')

MAXExpense=IntVar()

EMAXExpense= ttk.Entry(F2, textvariable=MAXExpense,font=(None,12))
EMAXExpense.grid(row=3, column=3,padx=5,pady=5,sticky='w')





BF2Add=ttk.Button(F2, text='Add Expense',command=Addexpense)
BF2Add.grid(row=3, column=1,padx=5,pady=5,sticky='w',ipadx=10,ipady=10)


BF2Delete=ttk.Button(F2, text='Delete Expense',command=Deleteexpense)
BF2Delete.grid(row=4, column=1,padx=5,pady=5,sticky='w',ipadx=10,ipady=10)


NFileButton=ttk.Button(F2, text='Create and Open New File',command=Open_New)
NFileButton.grid(row=0, column=2,padx=5,pady=5,sticky='w',ipadx=10,ipady=10)


FileButton=ttk.Button(F2, text='Open Existing File',command=Open_Existing)
FileButton.grid(row=0, column=3,padx=5,pady=5,sticky='w',ipadx=10,ipady=10)


MEButton=ttk.Button(F2, text='Month vs Expense',command=ME)
MEButton.grid(row=1, column=3,padx=5,pady=5,sticky='w',ipadx=10,ipady=10)


CatButton=ttk.Button(F2, text='Category-wise Expenditure',command=Show_Cat)
CatButton.grid(row=1, column=2,padx=5,pady=5,sticky='w',ipadx=10,ipady=10)


OButton=ttk.Button(F2, text='Overall Analysis',command=Show_Over)
OButton.grid(row=2, column=2,padx=5,pady=5,sticky='w',ipadx=10,ipady=10)


LimitButton=ttk.Button(F2, text='Check Status',command=Show_Status)
LimitButton.grid(row=4, column=3,padx=5,pady=5,sticky='w',ipadx=10,ipady=10)


clear_Button2=ttk.Button(F2, text='Clear',command=clear_expense_data)
clear_Button2.grid(row=5, column=3,padx=5,pady=5,sticky='w',ipadx=10,ipady=10)




#---TREE VIEW---
TVList2 =['Month','Category','Expense']
TV2 =ttk.Treeview(F2, column=TVList2, show='headings',height=14)
for i in TVList2:
    TV2.heading(i, text=i.title())
TV2.grid(row=5, column=0,padx=5,pady=5,sticky='w',columnspan=3)






GUI.mainloop()